"""
Exporta sinonimos_universal.json a Excel.
Genera: exportar/sinonimos_universal.xlsx
"""
import json, sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
SYNS_FILE  = BASE / 'sinonimos_universal.json'
HIST_FILE  = BASE / 'historial_universal.json'
OUT_FILE   = BASE / 'exportar' / 'sinonimos_universal.xlsx'

# ── Colores por origen ─────────────────────────────────────────────────────────
FILL = {
    'manual'    : PatternFill('solid', fgColor='C6EFCE'),   # verde claro
    'auto'      : PatternFill('solid', fgColor='DDEBF7'),   # azul claro
    'auto-fuzzy': PatternFill('solid', fgColor='FFEB9C'),   # amarillo
}
HEADER_FILL = PatternFill('solid', fgColor='2F4F8F')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
CELL_FONT   = Font(size=9)
THIN = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
)

def main():
    syns = json.loads(SYNS_FILE.read_text(encoding='utf-8'))
    hist = json.loads(HIST_FILE.read_text(encoding='utf-8')) if HIST_FILE.exists() else {}

    wb = openpyxl.Workbook()

    # ── Hoja 1: Sinónimos ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Sinónimos'

    headers = ['Proveedor ID', 'Especie', 'Variedad (factura)', 'Talla', 'SPB',
               'Grado', 'ID Artículo', 'Nombre Artículo VeraBuy', 'Origen']
    col_widths = [13, 14, 38, 7, 6, 8, 11, 42, 10]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = 'A2'

    # Ordenar por proveedor → especie → variedad
    rows = sorted(syns.values(), key=lambda v: (
        v.get('provider_id', 0), v.get('species', ''), v.get('variety', '')))

    for r, v in enumerate(rows, 2):
        origen = v.get('origen', '')
        fill   = FILL.get(origen, PatternFill())
        vals   = [
            v.get('provider_id'), v.get('species'), v.get('variety'),
            v.get('size') or '', v.get('stems_per_bunch') or '',
            v.get('grade', ''), v.get('articulo_id'),
            v.get('articulo_name'), origen,
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # ── Hoja 2: Historial ──────────────────────────────────────────────────────
    wh = wb.create_sheet('Historial')
    h2 = ['Fecha', 'Factura', 'Proveedor', 'PDF', 'Líneas', 'OK', 'Sin match', 'Total USD']
    cw = [17, 14, 26, 28, 7, 6, 10, 11]
    for c, (h, w) in enumerate(zip(h2, cw), 1):
        cell = wh.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
        wh.column_dimensions[get_column_letter(c)].width = w
    wh.row_dimensions[1].height = 18
    wh.freeze_panes = 'A2'

    hist_rows = sorted(hist.items(), key=lambda x: x[1].get('fecha',''), reverse=True)
    for r, (inv, d) in enumerate(hist_rows, 2):
        ok = d.get('ok', 0); tot = d.get('lineas', 0); fail = d.get('sin_match', 0)
        vals = [d.get('fecha'), inv, d.get('provider'), d.get('pdf'),
                tot, ok, fail, d.get('total_usd')]
        for c, val in enumerate(vals, 1):
            cell = wh.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT; cell.border = THIN
            cell.alignment = Alignment(vertical='center')
        # Colorear si hay sin_match
        if fail:
            for c in range(1, len(h2)+1):
                wh.cell(row=r, column=c).fill = PatternFill('solid', fgColor='FFE0E0')
    wh.auto_filter.ref = f'A1:{get_column_letter(len(h2))}1'

    OUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUT_FILE)
    print(f'Excel generado: {OUT_FILE}')
    print(f'  {len(syns)} sinónimos  |  {len(hist)} facturas en historial')

if __name__ == '__main__':
    main()
