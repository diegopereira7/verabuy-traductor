"""Auditoría de sinónimos — genera Excel para revisión y corrección.

Ejecutar:  python auditar_sinonimos.py [--output archivo.xlsx]

Genera un Excel con pestañas:
  1. Revisar       — TODOS los sinónimos con columna editable "Nuevo Art. ID"
  2. Sospechosos   — filtrado de los que probablemente están mal
  3. Duplicados    — misma variedad+talla con artículos diferentes
  4. Resumen       — estadísticas generales

Flujo:
  1. Ejecuta este script → genera Excel
  2. Abre el Excel, revisa pestaña "Revisar" o "Sospechosos"
  3. En columna "Nuevo Art. ID": escribe el ID correcto, o "BORRAR" para eliminar
  4. Guarda el Excel
  5. Ejecuta:  python importar_sinonimos.py auditoria_sinonimos.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from src.config import SQL_FILE, SYNS_FILE, PROVIDERS
from src.articulos import ArticulosLoader

sys.stdout.reconfigure(encoding='utf-8')

# ── Provider ID → name map ──────────────────────────────────────────────────
_PROV_NAMES = {v['id']: v['name'] for v in PROVIDERS.values()}


def _prov_name(pid: int) -> str:
    return _PROV_NAMES.get(pid, f'ID:{pid}')


def _normalize(s: str) -> str:
    return re.sub(r'\s+', ' ', s.strip().upper())


def _variety_from_article(nombre: str) -> str:
    """Extrae la variedad probable del nombre de artículo VeraBuy."""
    n = _normalize(nombre)
    n = re.sub(r'^(?:ROSA(?:\s+(?:EC|COL|GARDEN))?\s+|CLAVEL(?:\s+(?:FANCY|SELECT|ESTANDAR|SHORT|COL\s+\w+))?\s+|'
               r'MINI\s+CLAVEL(?:\s+COL)?\s+(?:FANCY|SELECT)?\s*|'
               r'ALSTROMERIA\s+(?:FANCY\s+|SELECT\s+|PERFECTION\s+)?|HYDRANGEA\s+(?:PREMIUM\s+)?|'
               r'ROSA\s+MONDIAL\s+)', '', n)
    n = re.sub(r'\s+\d+CM\s+\d+U(?:\s+[\w\s]+)?$', '', n)
    n = re.sub(r'\s+\d+CM\s*$', '', n)
    return n.strip()


_TRANSLATION = {
    'WHITE': 'BLANCO', 'RED': 'ROJO', 'PINK': 'ROSA', 'YELLOW': 'AMARILLO',
    'GREEN': 'VERDE', 'ORANGE': 'NARANJA', 'PURPLE': 'MORADO', 'BLUE': 'AZUL',
    'CREAM': 'CREMA', 'SALMON': 'SALMON', 'PEACH': 'MELOCOTON',
    'BICOLOR': 'BICOLOR', 'MIX': 'MIXTO', 'MIXED': 'MIXTO', 'ASSORTED': 'SURTIDO',
    'NEGRO': 'NEGRO', 'BLACK': 'NEGRO',
    'CAROUSEL': 'CARROUSEL', 'CARROUSEL': 'CARROUSEL',
    'NOVELTY': 'NOVEDADES', 'BURGUNDY': 'GRANATE',
}


def _similarity_score(variety: str, art_name: str) -> int:
    """Score 0-100 de cuánto se parece la variedad al nombre de artículo."""
    v = _normalize(variety)
    a = _normalize(art_name)
    if v in a:
        return 100
    art_var = _variety_from_article(a)
    if v == art_var:
        return 100
    if v in art_var or art_var in v:
        return 85
    # Translate color words
    v_translated = ' '.join(_TRANSLATION.get(w, w) for w in v.split())
    if v_translated in a or v_translated == art_var:
        return 95
    if art_var in v_translated or v_translated in art_var:
        return 85
    # Levenshtein-like for typos
    if art_var and abs(len(v) - len(art_var)) <= 2:
        common = sum(1 for a_c, b_c in zip(v, art_var) if a_c == b_c)
        ratio = common / max(len(v), len(art_var))
        if ratio >= 0.75:
            return int(ratio * 100)
    # Word overlap
    vw = set(v.split()) | set(v_translated.split())
    aw = set(art_var.split()) | set(a.split())
    if not vw:
        return 0
    v_words = set(v.split())
    overlap = len(vw & aw) / min(len(v_words), len(set(art_var.split()))) if art_var and v_words else 0
    return min(int(overlap * 85), 85)


def load_data():
    art = ArticulosLoader()
    art.load_from_sql(str(SQL_FILE))
    with open(SYNS_FILE, encoding='utf-8') as f:
        syns = json.load(f)
    return art, syns


def audit(art: ArticulosLoader, syns: dict):
    sospechosos = []
    todos = []
    variety_map = defaultdict(list)

    for key, syn in syns.items():
        aid = syn.get('articulo_id', 0)
        aname = syn.get('articulo_name', '')
        variety = syn.get('variety', '')
        species = syn.get('species', '')
        size = syn.get('size', 0)
        spb = syn.get('stems_per_bunch', 0)
        origen = syn.get('origen', '')
        pid = syn.get('provider_id', 0)
        raw = syn.get('raw', '')
        invoice = syn.get('invoice', '')

        exists = aid in art.articulos
        sim = _similarity_score(variety, aname) if aname else 0

        # Build display provider: "Nombre (ID:123)" for clarity
        prov_display = f"{_prov_name(pid)} (ID:{pid})"
        # Invoice: use raw to hint at source if no invoice
        invoice_display = invoice if invoice else ''

        entry = {
            'key': key, 'provider_id': pid, 'provider_name': _prov_name(pid),
            'provider_display': prov_display,
            'species': species, 'variety': variety, 'size': size, 'spb': spb,
            'articulo_id': aid, 'articulo_name': aname,
            'origen': origen, 'raw': raw[:120] if raw else '',
            'invoice': invoice_display,
            'similarity': sim, 'exists': exists,
        }
        todos.append(entry)

        if variety and size:
            variety_map[(_normalize(variety), size)].append(entry)

        # Flag suspicious — skip manually reviewed synonyms
        is_suspicious = False
        reason = []
        if origen in ('manual', 'manual-batch', 'revisado'):
            pass  # already reviewed by user, never flag
        else:
            if origen == 'auto-fuzzy' and sim < 80:
                reason.append(f'fuzzy baja similitud ({sim}%)')
            if origen == 'auto-fuzzy':
                reason.append('fuzzy')
                is_suspicious = True
            if sim < 50 and aid > 0:
                reason.append(f'similitud {sim}%')
                is_suspicious = True
            if not exists and aid > 0:
                reason.append('artículo eliminado')
                is_suspicious = True
            if species == 'ROSES' and aname and 'ROSA' not in aname.upper() and 'GARDEN' not in aname.upper():
                reason.append('especie≠artículo')
                is_suspicious = True
            if species == 'CARNATIONS' and aname and 'CLAVEL' not in aname.upper():
                reason.append('especie≠artículo')
                is_suspicious = True

        if is_suspicious:
            entry['reason'] = ' | '.join(reason)
            sospechosos.append(entry)

    duplicados = []
    for (var, sz), entries in variety_map.items():
        art_ids = set(e['articulo_id'] for e in entries)
        if len(art_ids) > 1:
            for e in entries:
                duplicados.append({**e, 'num_variants': len(art_ids)})

    return sospechosos, todos, duplicados


def generate_excel(sospechosos, todos, duplicados, output: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill('solid', fgColor='2F4F8F')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
    EDIT_FILL = PatternFill('solid', fgColor='FFFFCC')  # amarillo claro para editable
    EDIT_HEADER = PatternFill('solid', fgColor='FF8C00')  # naranja para columna editable
    RED = PatternFill('solid', fgColor='FFE0E0')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')
    GREEN = PatternFill('solid', fgColor='C6EFCE')
    THIN = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )

    wb = openpyxl.Workbook()

    def write_sheet(ws, headers, rows, widths, color_fn=None, edit_col=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = EDIT_HEADER if (edit_col and c == edit_col) else HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = THIN
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'A2'
        for r, row in enumerate(rows, 2):
            fill = color_fn(row) if color_fn else None
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = Font(size=9)
                cell.border = THIN
                if edit_col and c == edit_col:
                    cell.fill = EDIT_FILL
                elif fill:
                    cell.fill = fill
        if len(headers) <= 26:
            ws.auto_filter.ref = f'A1:{chr(64+len(headers))}1'

    # ��─ 1. Revisar (TODOS — editable) ────────────────────────────────────────
    ws = wb.active
    ws.title = 'Revisar'
    headers = [
        'Nuevo Art. ID\n(editar)',  # col 1 — EDITABLE
        'Proveedor', 'Factura', 'Variedad', 'Especie', 'Talla', 'SPB',
        'Art. ID actual', 'Artículo VeraBuy actual', 'Similitud', 'Origen',
        'Línea raw factura', 'Key sinónimo',
    ]
    widths = [14, 18, 18, 22, 14, 6, 5, 10, 42, 8, 13, 45, 35]
    rows = []
    for t in sorted(todos, key=lambda x: (x['provider_name'], x['species'], x['variety'])):
        rows.append([
            '',  # col 1: editable — vacío = sin cambio
            t['provider_display'], t['invoice'], t['variety'], t['species'],
            t['size'], t['spb'],
            t['articulo_id'], t['articulo_name'], t['similarity'], t['origen'],
            t['raw'], t['key'],
        ])
    write_sheet(ws, headers, rows, widths,
                color_fn=lambda r: RED if r[9] < 50 else (YELLOW if r[9] < 80 else GREEN),
                edit_col=1)

    # ── 2. Sospechosos (filtrado — editable) ─────────────────────────────────
    ws2 = wb.create_sheet('Sospechosos')
    headers2 = [
        'Nuevo Art. ID\n(editar)',  # col 1 — EDITABLE
        'Razón', 'Proveedor', 'Factura', 'Variedad', 'Especie', 'Talla',
        'Art. ID actual', 'Artículo VeraBuy actual', 'Similitud', 'Origen',
        'Línea raw factura', 'Key sinónimo',
    ]
    widths2 = [14, 28, 18, 18, 22, 14, 6, 10, 42, 8, 13, 45, 35]
    rows2 = []
    for s in sorted(sospechosos, key=lambda x: x['similarity']):
        rows2.append([
            '',
            s.get('reason', ''), s['provider_display'], s['invoice'],
            s['variety'], s['species'], s['size'],
            s['articulo_id'], s['articulo_name'], s['similarity'], s['origen'],
            s.get('raw', ''), s['key'],
        ])
    write_sheet(ws2, headers2, rows2, widths2,
                color_fn=lambda r: RED if r[9] < 50 else YELLOW,
                edit_col=1)

    # ── 3. Duplicados ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Duplicados')
    headers3 = ['Variedad', 'Talla', 'Proveedor', 'Art. ID', 'Artículo', 'Origen', '#Variantes']
    widths3 = [20, 6, 18, 10, 42, 14, 10]
    rows3 = []
    for d in sorted(duplicados, key=lambda x: (_normalize(x['variety']), x['size'])):
        rows3.append([
            d['variety'], d['size'], d['provider_display'],
            d['articulo_id'], d['articulo_name'], d['origen'], d['num_variants'],
        ])
    write_sheet(ws3, headers3, rows3, widths3, color_fn=lambda r: YELLOW)

    # ── 4. Resumen ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Resumen')
    origen_counts = Counter(t['origen'] for t in todos)
    sim_buckets = Counter()
    for t in todos:
        s = t['similarity']
        if s >= 90: sim_buckets['90-100 (excelente)'] += 1
        elif s >= 70: sim_buckets['70-89 (bueno)'] += 1
        elif s >= 50: sim_buckets['50-69 (dudoso)'] += 1
        else: sim_buckets['0-49 (sospechoso)'] += 1

    stats = [
        ['Total sinónimos', len(todos)],
        ['Sospechosos', len(sospechosos)],
        ['Duplicados (var+talla con múltiples artículos)', len(duplicados)],
        ['', ''],
        ['Por origen:', ''],
    ]
    for o, c in origen_counts.most_common():
        stats.append([f'  {o}', c])
    stats.extend([['', ''], ['Por similitud:', '']])
    for bucket in ['90-100 (excelente)', '70-89 (bueno)', '50-69 (dudoso)', '0-49 (sospechoso)']:
        stats.append([f'  {bucket}', sim_buckets.get(bucket, 0)])
    stats.extend([
        ['', ''],
        ['Instrucciones:', ''],
        ['  1. Revisa pestañas "Sospechosos" y "Revisar"', ''],
        ['  2. En columna "Nuevo Art. ID":', ''],
        ['     - Número (ej: 32779) para cambiar el artículo', ''],
        ['     - OK para marcar como revisado (no aparece más)', ''],
        ['     - BORRAR para eliminar el sinónimo', ''],
        ['     - Vacío = sin cambio', ''],
        ['  3. Guarda el Excel', ''],
        ['  4. Ejecuta: python importar_sinonimos.py auditoria_sinonimos.xlsx', ''],
    ])

    for r, (label, val) in enumerate(stats, 1):
        ws4.cell(row=r, column=1, value=label).font = Font(size=10, bold=bool(val == ''))
        ws4.cell(row=r, column=2, value=val).font = Font(size=10)
    ws4.column_dimensions['A'].width = 50
    ws4.column_dimensions['B'].width = 12

    output.parent.mkdir(exist_ok=True)
    wb.save(output)
    return output


def main():
    parser = argparse.ArgumentParser(description='Auditoría de sinónimos VeraBuy')
    parser.add_argument('--output', '-o', type=Path, default=Path('auditoria_sinonimos.xlsx'))
    args = parser.parse_args()

    print('Cargando datos...')
    art, syns = load_data()
    print(f'  {len(art.articulos)} artículos, {len(syns)} sinónimos')

    print('Analizando...')
    sospechosos, todos, duplicados = audit(art, syns)

    print(f'\nResultados:')
    print(f'  Total sinónimos: {len(todos)}')
    print(f'  Sospechosos:     {len(sospechosos)}')
    print(f'  Duplicados:      {len(duplicados)}')

    print(f'\nGenerando Excel...')
    out = generate_excel(sospechosos, todos, duplicados, args.output)
    print(f'Guardado en: {out}')
    print(f'\nEdita el Excel y luego ejecuta:')
    print(f'  python importar_sinonimos.py {out}')


if __name__ == '__main__':
    main()
