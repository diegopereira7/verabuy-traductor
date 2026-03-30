"""Procesador masivo de facturas PDF para VeraBuy.

Doble modo:
  CLI:  python batch_process.py carpeta/ --output resultado.xlsx
  Web:  python batch_process.py carpeta/ --batch-id abc123

Carga datos pesados (SQL + sinónimos) UNA VEZ y procesa todos los PDFs.
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
import traceback
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from src.config import SQL_FILE, SYNS_FILE, HIST_FILE, BASE_DIR
from src.pdf import detect_provider
from src.parsers import FORMAT_PARSERS
from src.articulos import ArticulosLoader
from src.sinonimos import SynonymStore
from src.matcher import Matcher, rescue_unparsed_lines, split_mixed_boxes
from src.historial import History

# Directorios de batch
STATUS_DIR  = BASE_DIR / 'batch_status'
RESULTS_DIR = BASE_DIR / 'batch_results'
UPLOADS_DIR = BASE_DIR / 'batch_uploads'


def _write_status(path: Path, data: dict):
    """Escritura atómica: escribe a .tmp y renombra para evitar lecturas parciales."""
    tmp = path.with_suffix('.tmp')
    tmp.write_text(json.dumps(data, ensure_ascii=False, default=str), encoding='utf-8')
    if path.exists():
        path.unlink()
    tmp.rename(path)


def _find_pdfs(folder: Path) -> list[Path]:
    """Busca todos los PDFs en la carpeta (recursivo)."""
    pdfs = sorted(folder.rglob('*.pdf'), key=lambda p: p.name.lower())
    return pdfs


def _process_single_pdf(
    pdf_path: Path,
    art: ArticulosLoader,
    syn: SynonymStore,
    matcher: Matcher,
) -> dict:
    """Procesa un PDF individual. Retorna dict con resultado o error."""
    pdata = detect_provider(str(pdf_path))
    if not pdata:
        return {
            'ok': False,
            'pdf': pdf_path.name,
            'error': 'Proveedor no reconocido',
        }

    pdata['pdf_path'] = str(pdf_path)  # para parsers que usan pdfplumber tables
    fmt = pdata.get('fmt', '')
    parser = FORMAT_PARSERS.get(fmt)
    if not parser:
        return {
            'ok': False,
            'pdf': pdf_path.name,
            'error': f'Sin parser para formato "{fmt}"',
        }

    header, lines = parser.parse(pdata['text'], pdata)
    lines = split_mixed_boxes(lines)
    rescued = rescue_unparsed_lines(pdata['text'], lines)
    lines = matcher.match_all(pdata['id'], lines)
    lines.extend(rescued)

    ok_count = sum(1 for l in lines if l.match_status == 'ok')
    no_parser = sum(1 for l in lines if l.match_status == 'sin_parser')

    return {
        'ok': True,
        'pdf': pdf_path.name,
        'header': {
            'invoice_number': header.invoice_number,
            'date':           header.date,
            'awb':            header.awb,
            'hawb':           getattr(header, 'hawb', ''),
            'provider_name':  header.provider_name,
            'provider_id':    header.provider_id,
            'total':          header.total,
        },
        'stats': {
            'total_lineas': len(lines),
            'ok':           ok_count,
            'sin_match':    len(lines) - ok_count - no_parser,
            'sin_parser':   no_parser,
        },
        'lines': [{
            'raw':             l.raw_description[:120],
            'species':         l.species,
            'variety':         l.variety,
            'grade':           l.grade,
            'size':            l.size,
            'stems_per_bunch': l.stems_per_bunch,
            'stems':           l.stems,
            'price_per_stem':  round(l.price_per_stem, 5),
            'line_total':      round(l.line_total, 2),
            'label':           l.label,
            'box_type':        l.box_type,
            'articulo_id':     l.articulo_id,
            'articulo_name':   l.articulo_name or '',
            'match_status':    l.match_status,
            'match_method':    l.match_method,
        } for l in lines],
    }


def _generate_excel(results: list[dict], output_path: Path):
    """Genera Excel consolidado con hoja Resumen + hoja Detalle."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill('solid', fgColor='2F4F8F')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
    CELL_FONT   = Font(size=9)
    THIN = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    GREEN  = PatternFill('solid', fgColor='C6EFCE')
    RED    = PatternFill('solid', fgColor='FFE0E0')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumen'
    headers = ['#', 'PDF', 'Proveedor', 'Factura', 'Fecha', 'AWB',
               'Líneas', 'OK', 'Sin Match', 'Total USD', 'Estado']
    widths  = [5, 30, 20, 15, 14, 18, 8, 6, 10, 12, 12]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

    for r, res in enumerate(results, 2):
        if res['ok']:
            h = res['header']
            s = res['stats']
            vals = [r - 1, res['pdf'], h['provider_name'], h['invoice_number'],
                    h['date'], h['awb'], s['total_lineas'], s['ok'],
                    s['sin_match'], h['total'], 'OK' if s['sin_match'] == 0 else 'PARCIAL']
            fill = GREEN if s['sin_match'] == 0 else YELLOW
        else:
            vals = [r - 1, res['pdf'], '', '', '', '', 0, 0, 0, 0, f"ERROR: {res['error']}"]
            fill = RED

        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical='center')

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    # ── Hoja 2: Detalle ──────────────────────────────────────────────────────
    wd = wb.create_sheet('Detalle')
    det_headers = ['PDF', 'Proveedor', 'Factura', 'Descripción', 'Especie',
                   'Variedad', 'Talla', 'SPB', 'Tallos', 'Precio/T', 'Total',
                   'ID Artículo', 'Nombre Artículo VeraBuy', 'Match']
    det_widths  = [22, 18, 14, 45, 14, 22, 7, 5, 7, 9, 10, 10, 42, 12]

    for c, (h, w) in enumerate(zip(det_headers, det_widths), 1):
        cell = wd.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN
        wd.column_dimensions[get_column_letter(c)].width = w
    wd.freeze_panes = 'A2'

    row = 2
    for res in results:
        if not res['ok']:
            continue
        h = res['header']
        for l in res['lines']:
            match_label = l['match_method'] if l['match_status'] == 'ok' else l['match_status'].upper()
            vals = [
                res['pdf'], h['provider_name'], h['invoice_number'],
                l['raw'], l['species'], l['variety'], l['size'] or '',
                l['stems_per_bunch'] or '', l['stems'] or '',
                l['price_per_stem'], l['line_total'],
                l['articulo_id'] or '', l['articulo_name'], match_label,
            ]
            fill = GREEN if l['match_status'] == 'ok' else (RED if l['match_status'] == 'sin_match' else YELLOW)
            for c, val in enumerate(vals, 1):
                cell = wd.cell(row=row, column=c, value=val)
                cell.font = CELL_FONT
                cell.fill = fill
                cell.border = THIN
                cell.alignment = Alignment(vertical='center')
            row += 1

    wd.auto_filter.ref = f'A1:{get_column_letter(len(det_headers))}1'

    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)


def run_batch(folder: Path, batch_id: str | None = None, output: Path | None = None):
    """Ejecuta el procesamiento masivo."""
    status_path = STATUS_DIR / f'{batch_id}.json' if batch_id else None
    is_web = batch_id is not None

    def update_status(data: dict):
        if status_path:
            data['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            _write_status(status_path, data)

    # Buscar PDFs
    pdfs = _find_pdfs(folder)
    total = len(pdfs)

    if total == 0:
        update_status({
            'estado': 'error',
            'progreso': 0, 'total': 0, 'porcentaje': 0,
            'error': 'No se encontraron archivos PDF en la carpeta',
        })
        if not is_web:
            print('ERROR: No se encontraron PDFs en la carpeta.')
        return

    # Cargar datos pesados UNA VEZ
    update_status({
        'estado': 'cargando_datos',
        'progreso': 0, 'total': total, 'porcentaje': 0,
        'actual': 'Cargando artículos y sinónimos...',
        'procesados_ok': 0, 'con_error': 0,
    })

    if not is_web:
        print(f'Encontrados {total} PDFs. Cargando datos...')

    art = ArticulosLoader()
    art.load_from_sql(str(SQL_FILE))
    syn = SynonymStore(str(SYNS_FILE))
    matcher = Matcher(art, syn)
    hist = History(str(HIST_FILE))

    if not is_web:
        print(f'  {len(art.articulos)} artículos, {syn.count()} sinónimos cargados.')

    # Patrones de nombre que NO son facturas de proveedor
    SKIP_PATTERNS = ['DUA', 'NYD', 'ALLIANCE', 'FULL', 'GUIA', 'PREALERT', 'REAL CARGA']

    def _should_skip(name: str) -> str | None:
        """Retorna el patrón detectado si el archivo debe omitirse, o None."""
        upper = name.upper()
        for pat in SKIP_PATTERNS:
            if pat in upper:
                return pat
        return None

    # Procesar cada PDF
    results = []
    errores = []
    omitidos = []
    ok_count = 0
    err_count = 0
    skip_count = 0

    for i, pdf in enumerate(pdfs):
        # Filtrar documentos que no son facturas
        skip_reason = _should_skip(pdf.name)
        if skip_reason:
            skip_count += 1
            omitidos.append({
                'pdf': pdf.name,
                'motivo': f'Omitido: documento no es factura ({skip_reason})',
            })
            if not is_web:
                print(f'  [{i+1}/{total}] {pdf.name}: OMITIDO — no es factura ({skip_reason})')
            continue

        update_status({
            'estado': 'procesando',
            'progreso': i, 'total': total,
            'porcentaje': round(i / total * 100),
            'actual': pdf.name,
            'procesados_ok': ok_count, 'con_error': err_count,
            'omitidos': skip_count,
        })

        try:
            result = _process_single_pdf(pdf, art, syn, matcher)
            results.append(result)

            if result['ok']:
                ok_count += 1
                h = result['header']
                s = result['stats']
                # Registrar en historial
                hist.add(
                    h['invoice_number'], pdf.name, h['provider_name'],
                    h['total'], s['total_lineas'], s['ok'], s['sin_match'],
                )
                if not is_web:
                    status = 'OK' if s['sin_match'] == 0 else f'{s["sin_match"]} sin match'
                    print(f'  [{i+1}/{total}] {pdf.name}: {s["ok"]}/{s["total_lineas"]} — {status}')
            else:
                err_count += 1
                errores.append({'pdf': pdf.name, 'error': result['error']})
                if not is_web:
                    print(f'  [{i+1}/{total}] {pdf.name}: ERROR — {result["error"]}')

        except Exception as exc:
            err_count += 1
            tb = traceback.format_exc()
            errores.append({'pdf': pdf.name, 'error': str(exc), 'traceback': tb})
            results.append({'ok': False, 'pdf': pdf.name, 'error': str(exc)})
            if not is_web:
                print(f'  [{i+1}/{total}] {pdf.name}: EXCEPCIÓN — {exc}')

    # Guardar historial
    hist.save()

    # Auto-aprendizaje: intentar aprender de los PDFs que fallaron
    pdfs_sin_parser = [
        {'path': str(pdf), 'text': r.get('_text', ''), 'fingerprint': None}
        for pdf, r in zip(pdfs, results)
        if not r['ok'] and 'no reconocido' in r.get('error', '').lower()
    ]
    if len(pdfs_sin_parser) >= 2:
        try:
            from src.learner import aprender_de_batch
            update_status({
                'estado': 'aprendiendo',
                'progreso': total, 'total': total, 'porcentaje': 98,
                'actual': f'Auto-aprendizaje: {len(pdfs_sin_parser)} PDFs sin parser...',
                'procesados_ok': ok_count, 'con_error': err_count,
            })
            learned = aprender_de_batch(pdfs_sin_parser)
            if learned:
                # Reprocesar los PDFs que ahora tienen parser
                for lr in learned:
                    if lr['decision'] in ('VERDE', 'AMARILLO'):
                        for pdf_path in lr['pdfs']:
                            pdf = Path(pdf_path)
                            try:
                                reresult = _process_single_pdf(pdf, art, syn, matcher)
                                if reresult['ok']:
                                    # Reemplazar el resultado de error por el nuevo
                                    for j, r in enumerate(results):
                                        if r['pdf'] == pdf.name and not r['ok']:
                                            results[j] = reresult
                                            ok_count += 1
                                            err_count -= 1
                                            break
                            except Exception:
                                pass
                if not is_web:
                    print(f'\n  Auto-aprendizaje: {len(learned)} parsers generados')
        except Exception as exc:
            if not is_web:
                print(f'\n  Auto-aprendizaje falló: {exc}')

    # Generar Excel
    update_status({
        'estado': 'generando_excel',
        'progreso': total, 'total': total, 'porcentaje': 99,
        'actual': 'Generando Excel consolidado...',
        'procesados_ok': ok_count, 'con_error': err_count,
    })

    if output:
        excel_path = output
    elif batch_id:
        excel_path = RESULTS_DIR / f'{batch_id}.xlsx'
    else:
        excel_path = Path('batch_resultado.xlsx')

    _generate_excel(results, excel_path)

    if not is_web:
        print(f'\nExcel generado: {excel_path}')

    # Estadísticas totales
    total_lineas = sum(r['stats']['total_lineas'] for r in results if r['ok'])
    total_ok     = sum(r['stats']['ok'] for r in results if r['ok'])
    total_fail   = sum(r['stats']['sin_match'] for r in results if r['ok'])
    total_usd    = sum(r['header']['total'] for r in results if r['ok'])

    resumen = {
        'total_facturas':  total,
        'procesadas_ok':   ok_count,
        'con_error':       err_count,
        'omitidos':        skip_count,
        'total_lineas':    total_lineas,
        'total_ok':        total_ok,
        'total_sin_match': total_fail,
        'total_usd':       round(total_usd, 2),
    }

    # Estado final
    final_status = {
        'estado': 'completado',
        'progreso': total, 'total': total, 'porcentaje': 100,
        'actual': '',
        'procesados_ok': ok_count, 'con_error': err_count, 'omitidos': skip_count,
        'resumen': resumen,
        'omitidos_detalle': omitidos,
        'resultados': [{
            'pdf': r['pdf'],
            'ok': r['ok'],
            'provider': r['header']['provider_name'] if r['ok'] else '',
            'provider_id': r['header']['provider_id'] if r['ok'] else 0,
            'invoice': r['header']['invoice_number'] if r['ok'] else '',
            'date': r['header']['date'] if r['ok'] else '',
            'lineas': r['stats']['total_lineas'] if r['ok'] else 0,
            'ok_count': r['stats']['ok'] if r['ok'] else 0,
            'sin_match': r['stats']['sin_match'] if r['ok'] else 0,
            'total_usd': r['header']['total'] if r['ok'] else 0,
            'error': r.get('error', ''),
            'lines': r.get('lines', []) if r['ok'] else [],
        } for r in results],
        'errores': errores,
        'excel': excel_path.name,
    }
    update_status(final_status)

    if not is_web:
        omit_msg = f', {skip_count} omitidos' if skip_count else ''
        print(f'\nResumen: {ok_count} OK, {err_count} errores{omit_msg}, '
              f'{total_lineas} líneas ({total_ok} match, {total_fail} sin match)')
        print(f'Total USD: ${total_usd:,.2f}')

    # Limpiar carpeta temporal (solo modo web)
    if is_web and folder.parent == UPLOADS_DIR:
        shutil.rmtree(folder, ignore_errors=True)


def main():
    parser = argparse.ArgumentParser(description='Procesador masivo de facturas PDF VeraBuy')
    parser.add_argument('folder', type=Path, help='Carpeta con PDFs a procesar')
    parser.add_argument('--output', '-o', type=Path, default=None,
                        help='Ruta del Excel de salida (modo CLI)')
    parser.add_argument('--batch-id', type=str, default=None,
                        help='ID de batch (modo web, escribe status JSON)')
    args = parser.parse_args()

    if not args.folder.is_dir():
        print(f'ERROR: {args.folder} no es un directorio válido.')
        sys.exit(1)

    run_batch(args.folder, batch_id=args.batch_id, output=args.output)


if __name__ == '__main__':
    main()
