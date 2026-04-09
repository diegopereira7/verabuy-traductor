"""Importa correcciones de sinónimos desde el Excel de auditoría.

Ejecutar:  python importar_sinonimos.py auditoria_sinonimos.xlsx

Lee la columna "Nuevo Art. ID" de las pestañas "Revisar" y "Sospechosos":
  - Número → cambia el articulo_id (y busca el nombre en el catálogo)
  - "BORRAR" → elimina el sinónimo
  - Vacío → sin cambio

Hace backup antes de modificar sinonimos_universal.json.
"""
from __future__ import annotations

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from src.config import SQL_FILE, SYNS_FILE, FILE_ENCODING
from src.articulos import ArticulosLoader

sys.stdout.reconfigure(encoding='utf-8')


def load_excel(path: Path) -> list[dict]:
    """Lee las correcciones de ambas pestañas editables."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    changes = []

    for sheet_name in ('Revisar', 'Sospechosos'):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Find column indices from header row
        headers = [str(c.value or '').strip() for c in ws[1]]

        # Column positions differ per sheet
        new_id_col = None
        key_col = None
        for i, h in enumerate(headers):
            if 'Nuevo Art' in h:
                new_id_col = i
            if 'Key sin' in h:
                key_col = i

        if new_id_col is None or key_col is None:
            print(f'  Pestaña "{sheet_name}": columnas no encontradas, saltando')
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            new_val = row[new_id_col]
            key = row[key_col]
            if not key or new_val is None or str(new_val).strip() == '':
                continue
            new_str = str(new_val).strip().upper()
            changes.append({'key': str(key).strip(), 'action': new_str, 'sheet': sheet_name})

    wb.close()
    # Deduplicate: if same key appears in both sheets, last one wins
    seen = {}
    for c in changes:
        seen[c['key']] = c
    return list(seen.values())


def apply_changes(changes: list[dict], syns: dict, art: ArticulosLoader) -> dict:
    """Aplica los cambios al diccionario de sinónimos."""
    deleted = 0
    updated = 0
    errors = []

    for change in changes:
        key = change['key']
        action = change['action']

        if key not in syns:
            errors.append(f'Key no encontrada: {key}')
            continue

        if action == 'BORRAR':
            del syns[key]
            deleted += 1
            print(f'  BORRADO: {key}')
            continue

        if action == 'OK':
            syns[key]['origen'] = 'revisado'
            updated += 1
            print(f'  REVISADO OK: {key}')
            continue

        # Try to parse as article ID
        try:
            new_id = int(float(action))  # float() handles "32779.0" from Excel
        except (ValueError, TypeError):
            errors.append(f'Valor inválido para {key}: {action!r} (esperado número o BORRAR)')
            continue

        # Validate article exists
        if new_id not in art.articulos:
            errors.append(f'Art. ID {new_id} no existe en catálogo (key: {key})')
            continue

        art_info = art.articulos[new_id]
        new_name = art_info.get('nombre', '')
        old_id = syns[key].get('articulo_id', 0)
        old_name = syns[key].get('articulo_name', '')

        syns[key]['articulo_id'] = new_id
        syns[key]['articulo_name'] = new_name
        syns[key]['origen'] = 'manual'
        updated += 1
        print(f'  ACTUALIZADO: {key}')
        print(f'    {old_id} ({old_name[:40]})')
        print(f'    → {new_id} ({new_name[:40]})')

    return {'deleted': deleted, 'updated': updated, 'errors': errors}


def main():
    if len(sys.argv) < 2:
        print('Uso: python importar_sinonimos.py <archivo.xlsx>')
        print('  Lee correcciones del Excel de auditoría y actualiza sinonimos_universal.json')
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f'ERROR: {excel_path} no existe')
        sys.exit(1)

    # Load
    print(f'Leyendo correcciones de {excel_path}...')
    changes = load_excel(excel_path)
    if not changes:
        print('No hay cambios para aplicar (columna "Nuevo Art. ID" está vacía en todas las filas).')
        sys.exit(0)
    print(f'  {len(changes)} correcciones encontradas')

    print('Cargando catálogo...')
    art = ArticulosLoader()
    art.load_from_sql(str(SQL_FILE))

    print('Cargando sinónimos...')
    with open(SYNS_FILE, encoding=FILE_ENCODING) as f:
        syns = json.load(f)
    print(f'  {len(syns)} sinónimos cargados')

    # Backup
    backup = Path(f'sinonimos_backup_{datetime.now():%Y%m%d_%H%M%S}.json')
    shutil.copy2(SYNS_FILE, backup)
    print(f'  Backup: {backup}')

    # Apply
    print(f'\nAplicando {len(changes)} cambios...')
    result = apply_changes(changes, syns, art)

    # Save
    if result['updated'] > 0 or result['deleted'] > 0:
        with open(SYNS_FILE, 'w', encoding=FILE_ENCODING) as f:
            json.dump(syns, f, ensure_ascii=False, indent=2)
        print(f'\nGuardado en {SYNS_FILE}')
    else:
        print(f'\nSin cambios válidos — archivo no modificado.')

    # Summary
    print(f'\nResumen:')
    print(f'  Actualizados: {result["updated"]}')
    print(f'  Borrados:     {result["deleted"]}')
    if result['errors']:
        print(f'  Errores:      {len(result["errors"])}')
        for e in result['errors']:
            print(f'    ⚠ {e}')


if __name__ == '__main__':
    main()
