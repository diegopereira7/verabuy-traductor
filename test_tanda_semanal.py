"""Test de procesamiento masivo con una tanda semanal de facturas.

Uso:
    python test_tanda_semanal.py facturas_test/
    python test_tanda_semanal.py facturas_test/SEMANA\ 7/

Procesa todos los PDFs de la carpeta (recursivo), genera informe
resumen y opcionalmente un Excel con todas las líneas.
"""
from __future__ import annotations

import sys
import traceback
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from src.config import SQL_FILE, SYNS_FILE, HIST_FILE
from src.pdf import detect_provider
from src.parsers import FORMAT_PARSERS
from src.articulos import ArticulosLoader
from src.sinonimos import SynonymStore
from src.matcher import Matcher, rescue_unparsed_lines, split_mixed_boxes, reclassify_assorted
from src.historial import History


# Patrones de nombre que NO son facturas de proveedor
SKIP_PATTERNS = [
    'DUA', 'NYD', 'ALLIANCE', 'FULL', 'GUIA', 'PREALERT', 'PRE ALERT',
    'REAL CARGA', 'SAFTEC', 'EXCELLENT', 'EXCELLENTE',
    'BTOS', 'PARTE', 'CORRECTA', 'LINDA', 'JORGE', 'SLU',
]


def should_skip(name: str) -> str | None:
    upper = name.upper()
    for pat in SKIP_PATTERNS:
        if pat in upper:
            return pat
    return None


def process_pdf(pdf_path: Path, art, syn, matcher) -> dict:
    """Procesa un PDF. Retorna dict con resultado o error."""
    pdata = detect_provider(str(pdf_path))
    if not pdata:
        return {'ok': False, 'pdf': pdf_path.name, 'error': 'Proveedor no reconocido',
                'lines': [], 'header': {}, 'stats': {}}

    pdata['pdf_path'] = str(pdf_path)
    fmt = pdata.get('fmt', '')
    parser = FORMAT_PARSERS.get(fmt)
    if not parser:
        return {'ok': False, 'pdf': pdf_path.name,
                'error': f'Sin parser para formato "{fmt}"',
                'lines': [], 'header': {}, 'stats': {}}

    header, lines = parser.parse(pdata['text'], pdata)
    lines = split_mixed_boxes(lines)
    rescued = rescue_unparsed_lines(pdata['text'], lines)
    lines = matcher.match_all(pdata['id'], lines, invoice=header.invoice_number)
    lines = reclassify_assorted(lines)
    lines.extend(rescued)

    ok = sum(1 for l in lines if l.match_status == 'ok')
    no_parser = sum(1 for l in lines if l.match_status == 'sin_parser')
    sin_match = len(lines) - ok - no_parser

    return {
        'ok': True,
        'pdf': pdf_path.name,
        'provider': pdata.get('name', '?'),
        'provider_id': pdata.get('id', 0),
        'header': {
            'invoice_number': header.invoice_number,
            'date': header.date,
            'awb': header.awb,
            'provider_name': header.provider_name,
            'total': header.total,
        },
        'stats': {'total': len(lines), 'ok': ok, 'sin_match': sin_match, 'sin_parser': no_parser},
        'lines': lines,
    }


def generate_excel(all_results: list[dict], output_path: Path):
    """Genera Excel con hoja Resumen + Detalle."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # Hoja Resumen
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['#', 'PDF', 'Proveedor', 'Factura', 'AWB', 'Lineas', 'OK',
               'Sin Match', 'Sin Parser', '% OK', 'Total USD'])
    for i, r in enumerate(all_results, 1):
        if not r['ok']:
            ws.append([i, r['pdf'], '', '', '', 0, 0, 0, 0, '', f"ERROR: {r['error']}"])
            continue
        s = r['stats']
        pct = f"{s['ok']/s['total']*100:.0f}%" if s['total'] else '0%'
        h = r['header']
        ws.append([i, r['pdf'], r.get('provider', ''), h.get('invoice_number', ''),
                   h.get('awb', ''), s['total'], s['ok'], s['sin_match'],
                   s['sin_parser'], pct, h.get('total', 0)])

    # Hoja Detalle
    wd = wb.create_sheet('Detalle')
    wd.append(['PDF', 'Proveedor', 'Especie', 'Variedad', 'Talla', 'SPB', 'Tallos',
               'Total', 'Match Status', 'Match Method', 'Articulo ID', 'Articulo Name'])
    for r in all_results:
        if not r['ok']:
            continue
        for l in r['lines']:
            wd.append([r['pdf'], r.get('provider', ''), l.species, l.variety,
                       l.size, l.stems_per_bunch, l.stems, round(l.line_total, 2),
                       l.match_status, l.match_method, l.articulo_id,
                       l.articulo_name or ''])

    wb.save(output_path)


def print_report(all_results: list[dict], skipped: int, errors: list):
    """Imprime informe resumen."""
    total_pdfs = len(all_results)
    ok_pdfs = sum(1 for r in all_results if r['ok'])
    err_pdfs = len(errors)

    total_lines = sum(r['stats']['total'] for r in all_results if r['ok'])
    ok_lines = sum(r['stats']['ok'] for r in all_results if r['ok'])
    sm_lines = sum(r['stats']['sin_match'] for r in all_results if r['ok'])
    sp_lines = sum(r['stats']['sin_parser'] for r in all_results if r['ok'])
    pct = ok_lines / total_lines * 100 if total_lines else 0

    print("\n" + "=" * 80)
    print("  INFORME DE PROCESAMIENTO — TANDA SEMANAL")
    print("=" * 80)
    print(f"  PDFs procesados:       {ok_pdfs}")
    print(f"  PDFs con error:        {err_pdfs}")
    print(f"  PDFs omitidos:         {skipped}")
    print(f"  Total lineas:          {total_lines}")
    print(f"  Lineas OK:             {ok_lines}")
    print(f"  Lineas SIN_MATCH:      {sm_lines}")
    print(f"  Lineas SIN_PARSER:     {sp_lines}")
    print(f"  % resolucion global:   {pct:.1f}%")

    # Desglose por proveedor
    prov_stats = defaultdict(lambda: {'pdfs': 0, 'lines': 0, 'ok': 0, 'sm': 0, 'sp': 0})
    for r in all_results:
        if not r['ok']:
            continue
        p = r.get('provider', '?')
        s = r['stats']
        prov_stats[p]['pdfs'] += 1
        prov_stats[p]['lines'] += s['total']
        prov_stats[p]['ok'] += s['ok']
        prov_stats[p]['sm'] += s['sin_match']
        prov_stats[p]['sp'] += s['sin_parser']

    print("\n  --- POR PROVEEDOR ---")
    print(f"  {'Proveedor':30s} {'PDFs':>5s} {'Lineas':>7s} {'OK':>5s} {'SinM':>6s} {'SinP':>6s} {'%OK':>6s}")
    print("  " + "-" * 72)
    for prov in sorted(prov_stats, key=lambda p: -prov_stats[p]['lines']):
        d = prov_stats[prov]
        ppct = d['ok'] / d['lines'] * 100 if d['lines'] else 0
        print(f"  {prov:30s} {d['pdfs']:5d} {d['lines']:7d} {d['ok']:5d} {d['sm']:6d} {d['sp']:6d} {ppct:5.1f}%")

    # Top variedades sin resolver
    var_counter = Counter()
    for r in all_results:
        if not r['ok']:
            continue
        for l in r['lines']:
            if l.match_status == 'sin_match':
                key = f"{r.get('provider', '?')}|{l.species}|{l.variety}|{l.size}"
                var_counter[key] += 1

    if var_counter:
        print("\n  --- TOP 20 VARIEDADES SIN RESOLVER ---")
        for key, cnt in var_counter.most_common(20):
            parts = key.split('|')
            print(f"  {cnt:4d}x  {parts[0]:25s} {parts[1]:15s} {parts[2]:25s} {parts[3]}cm")

    # Match methods
    method_counter = Counter()
    syn_existing = 0
    syn_new = 0
    for r in all_results:
        if not r['ok']:
            continue
        for l in r['lines']:
            if l.match_status == 'ok':
                method_counter[l.match_method] += 1
                if 'sinónimo' in l.match_method or l.match_method == 'sinónimo':
                    syn_existing += 1
                else:
                    syn_new += 1

    print("\n  --- METODOS DE MATCH ---")
    for method, cnt in method_counter.most_common(15):
        print(f"  {cnt:5d}  {method}")
    print(f"\n  Sinonimos existentes usados: {syn_existing}")
    print(f"  Matches automaticos nuevos:  {syn_new}")

    # Errors
    if errors:
        print("\n  --- ERRORES ---")
        for e in errors[:10]:
            print(f"  {e['pdf']}: {e['error']}")

    print("=" * 80)


def main():
    if len(sys.argv) < 2:
        print("Uso: python test_tanda_semanal.py <carpeta_pdfs>")
        print("Ejemplo: python test_tanda_semanal.py facturas_test/SEMANA\\ 7/")
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f"Error: '{folder}' no es una carpeta valida")
        sys.exit(1)

    print(f"=== TEST TANDA SEMANAL ===")
    print(f"Carpeta: {folder}")

    # Buscar PDFs
    pdfs = sorted(folder.rglob('*.pdf'), key=lambda p: p.name.lower())
    print(f"PDFs encontrados: {len(pdfs)}")

    # Cargar datos
    print("Cargando articulos y sinonimos...")
    art = ArticulosLoader()
    art.load_from_sql(str(SQL_FILE))
    syn = SynonymStore(str(SYNS_FILE))
    matcher = Matcher(art, syn)
    hist = History(str(HIST_FILE))
    print(f"  {len(art.articulos)} articulos, {syn.count()} sinonimos")

    # Procesar
    all_results = []
    errors = []
    skipped = 0

    for i, pdf in enumerate(pdfs):
        skip_reason = should_skip(pdf.name)
        if skip_reason:
            skipped += 1
            print(f"  [{i+1}/{len(pdfs)}] {pdf.name}: OMITIDO ({skip_reason})")
            continue

        try:
            result = process_pdf(pdf, art, syn, matcher)
            all_results.append(result)

            if result['ok']:
                s = result['stats']
                h = result['header']
                status = 'OK' if s['sin_match'] == 0 and s['sin_parser'] == 0 else \
                         f"{s['sin_match']} sin_match, {s['sin_parser']} sin_parser"
                print(f"  [{i+1}/{len(pdfs)}] {pdf.name}: {s['ok']}/{s['total']} -- {status}")
                # Registrar en historial
                hist.add(h['invoice_number'], pdf.name, h['provider_name'],
                         h['total'], s['total'], s['ok'], s['sin_match'])
            else:
                errors.append({'pdf': pdf.name, 'error': result['error']})
                print(f"  [{i+1}/{len(pdfs)}] {pdf.name}: ERROR -- {result['error']}")

        except Exception as exc:
            errors.append({'pdf': pdf.name, 'error': str(exc)})
            print(f"  [{i+1}/{len(pdfs)}] {pdf.name}: EXCEPCION -- {exc}")
            traceback.print_exc()

    # Informe
    print_report(all_results, skipped, errors)

    # Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    excel_path = Path(f"tanda_resultado_{timestamp}.xlsx")
    generate_excel(all_results, excel_path)
    print(f"\nExcel generado: {excel_path}")


if __name__ == '__main__':
    main()
