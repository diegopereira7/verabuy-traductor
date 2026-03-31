"""
Matching masivo de líneas pendientes contra artículos VeraBuy.

Replica el criterio manual del usuario:
  1. Buscar variedad en artículos (inglés y español)
  2. Filtrar por talla
  3. Buscar por marca/farm del proveedor (PRIORIDAD MÁXIMA)
  4. Buscar por campo proveedor
  5. Fallback a genérico EC/COL
  6. Artículo genérico sin talla
  7. Sin match → vacío
"""
from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

# --- Configuración ---
SQL_FILE = Path("articulos (3).sql")
EXCEL_FILE = Path("lineas_pendientes_2026.xlsx")
SYNS_FILE = Path("sinonimos_universal.json")
OUTPUT_EXCEL = Path("lineas_pendientes_2026_matched.xlsx")
OUTPUT_SYNS_ALTA = Path("sinonimos_matching_alta_media.json")
OUTPUT_SYNS_BAJA = Path("sinonimos_matching_baja.json")

# Provider name → known brand suffixes in articles
# Built from manual analysis + SQL data
PROVIDER_BRAND_MAP = {
    "Cantiza Flores":       ["CANTIZA", "CZ"],
    "Valthomig":            ["CANTIZA", "CZ"],   # same fmt as cantiza
    "Inversiones del Neusa":["ALUNA"],
    "Brissas":              ["BRISSAS"],
    "La Alegria Farm":      ["LA ALEGRIA", "ALEGRIA"],
    "Ceresfarms":           ["CERES"],
    "Florsani":             ["FLORSANI"],
    "Uma Flowers":          ["VIOLETA"],
    "Stampsybox":           ["STAMPSY"],
    "Fiorentina Flowers":   ["FIORENTINA"],
    "Benchmark Growers":    ["GOLDEN"],
    "Olimpo Flowers":       ["OLIMPO"],
    "Luxus Blumen":         ["LUXUS"],
    "Agrivaldani":          ["AGRIVALDANI"],
    "Colibri Flowers":      ["COLIBRI"],
    "Latin Flowers":        ["LATIN"],
    "Multiflora":           ["MULTIFLORA"],
    "Life Flowers":         ["LIFEFLOWERS"],
    "Floricola La Rosaleda":["ROSALEDA"],
    "Maxiflores":           ["MAXI", "MAXIFLORES"],
    "Mystic Flowers":       ["MYSTIC"],
    "Monterosas Farms":     ["MONTEROSAS"],
    "EQR USA":              ["EQR"],
    "Sayonara":             ["SAYONARA"],
    "Rosely Flowers":       ["ROSELY"],
    "Prestige Roses":       ["PRESTIGE"],
    "Turflor":              ["TURFLOR"],
    "Daflor":               ["DAFLOR"],
    "Valle Verde Farms":    ["VERDE", "VALLEVERDE"],
    "Floraroma":            ["FLORAROMA", "AROMA"],
    "Condor Andino":        ["CONDOR"],
    "Starflowers (Malima)": ["MALIMA", "STARFLOWERS"],
    "Tessa Corp":           ["TESSA"],
    "Verdes La Estacion":   ["VERDES", "ESTACION"],
    "Secore Floral":        ["SECORE"],
    "Ecoflor Groupchile":   ["ECOFLOR"],
    "Flores y Frutas Florifrut": ["FLORIFRUT"],
    "Tierra Verde":         ["TIERRAVERDE"],
    "Laila Flowers":        ["LAILA"],
    "Flores de la Hacienda":["HACIENDA"],
    "Rosadex":              ["ROSADEX"],
    "Cananvalley Flowers":  ["CANANVALLE"],
    "Plantaciones Plantreb":["TREBOL"],
    "Naranjo Roses":        ["NARANJO"],
    "Flores de Aposentos":  ["APOSENTOS"],
    "Utopia Farms":         ["ESMERALDA", "UTOPIA"],
    "Calinama (Native)":    ["NATIVE", "CALINAMA"],
    "Unique Flowers":       ["UNIQUE"],
    "Bosque Flowers":       ["BOSQUE"],
    "GardaExport":          ["GARDA"],
    "Floricola Latinafarms":["LATINA"],
    "Agricola Circasia":    ["CIRCASIA"],
    "Vuelven":              ["VUELVEN"],
    "Flores Milonga":       ["MILONGA"],
    "Comercializadora Almanti": ["MUCH", "ALMANTI"],
    "Premium Flowers Boyacá": ["PREMIUM"],
    "Domenica (Valencia Pozo)": ["DOMENICA"],
    "Invos Flowers":        ["INVOS"],
    "Meaflos":              ["MEAFLOS"],
    "Heraflor":             ["HERAFLOR"],
    "Infinity Trading":     ["INFINITY"],
    "Flores El Progreso":   ["PROGRESO"],
    "C.I. Flores Colon":    ["COLON"],
    "Agrícola Aguablanca":  ["AGUABLANCA"],
    "Success Flowers":      ["SUCCESS"],
}

# Traducciones variedad/color inglés → español
TRANSLATIONS = {
    # Species
    "GYPSOPHILA": "PANICULATA",
    "GYPSO": "PANICULATA",
    "GYPSO.": "PANICULATA",
    # Colors
    "WHITE": "BLANCO",
    "RED": "ROJO",
    "YELLOW": "AMARILLO",
    "PINK": "ROSA",
    "BLUE": "AZUL",
    "ORANGE": "NARANJA",
    "GREEN": "VERDE",
    "PURPLE": "MORADO",
    "LIGHT": "CLARO",
    "DARK": "OSCURO",
    "TINTED": "TEÑIDA",
    "PRESERVED": "PRESERVADA",
    "NATURAL": "BLANCO",  # Natural gypsophila = blanco
    "OVERTIME": "OVERTIME",
    "OVER TIME": "OVERTIME",
}

# Known variety spelling variations (factura → artículo)
VARIETY_ALIASES = {
    "JESSIKA": "JESSICA",
    "TOFFE": "TOFFEE",
    "SWEETNESS": "SWEET NESS",
    "SWEET NESS": "SWEETNESS",
    "DEEP PURPLE": "DEEP PURPLE",
    "PINK MONDIAL": "PINK MONDIAL",
    "GARDEN ROSE APPLE JACK": "APPLE JACK",
    "GARDEN ROSE": "",
}


def normalize(s: str) -> str:
    """Normaliza texto: upper, strip tildes, strip extra spaces."""
    s = s.upper().strip()
    # Remove accents
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def load_articles(sql_path: Path) -> dict:
    """Load articles from SQL dump. Returns dict id -> {nombre, id_proveedor}."""
    articles = {}

    def split_sql(s):
        parts, cur, inq, esc = [], "", False, False
        for ch in s:
            if esc:
                cur += ch; esc = False; continue
            if ch == "\\":
                cur += ch; esc = True; continue
            if ch == "'" and not inq:
                inq = True; cur += ch; continue
            if ch == "'" and inq:
                inq = False; cur += ch; continue
            if ch == "," and not inq:
                parts.append(cur); cur = ""; continue
            cur += ch
        if cur:
            parts.append(cur)
        return parts

    def cs(v):
        v = v.strip()
        if v == "NULL":
            return ""
        if v.startswith("'") and v.endswith("'"):
            return v[1:-1].replace("\\'", "'")
        return v

    def ci(v):
        v = v.strip()
        return int(v) if v.replace("-", "").isdigit() else 0

    with open(sql_path, "r", encoding="utf-8") as f:
        for raw in f:
            ln = raw.strip()
            if not ln.startswith("("):
                continue
            ln = ln.rstrip(",;")
            inner = ln[1:].rstrip(")")
            flds = split_sql(inner)
            if len(flds) < 10:
                continue
            try:
                articles[ci(flds[0])] = {
                    "id": ci(flds[0]),
                    "nombre": cs(flds[8]),
                    "id_proveedor": ci(flds[3]),
                    "familia": cs(flds[9]),
                }
            except Exception:
                continue

    return articles


def build_indices(articles: dict):
    """Build search indices from articles."""
    # Index: normalized variety → list of (article_id, nombre, size, spb, brand)
    variety_index = defaultdict(list)  # variety → [(id, nombre, size, spb, brand_suffix)]
    name_index = {}  # nombre_upper → id
    brand_index = defaultdict(list)  # brand → [(id, nombre)]
    provider_index = defaultdict(list)  # id_proveedor → [(id, nombre)]

    for art_id, art in articles.items():
        nombre = art["nombre"]
        if not nombre:
            continue

        n_upper = nombre.upper().strip()
        name_index[n_upper] = art_id

        # Index by provider
        if art["id_proveedor"]:
            provider_index[art["id_proveedor"]].append((art_id, nombre))

        # Parse rose articles: ROSA [EC|COL] VARIETY SIZECM SPBU [BRAND]
        m = re.match(
            r"ROSA\s+(?:EC\s+|COL\s+)?(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()

            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "ROSES",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse tinted roses: ROSA TEÑIDA COLOR SIZECM SPBU [BRAND]
        m = re.match(
            r"ROSA\s+TE[ÑN]IDA\s+(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            color = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety = f"TEÑIDA {color}"

            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "ROSES",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse garden roses: ROSA EC VARIETY SIZECM SPBU
        m = re.match(
            r"ROSA\s+(?:LA\s+)?(?:GARDEN\s+)?(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "ROSES",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse paniculata/gypsophila: PANICULATA VARIETY WEIGHT SIZE BRAND
        m = re.match(
            r"PANICULATA\s+(.+?)(?:\s+(\d+)GR)?(?:\s+(\d+)U)?(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            brand = (m.group(4) or "").strip()
            variety_index[normalize(f"PANICULATA {variety}")].append({
                "id": art_id, "nombre": nombre, "size": 0,
                "spb": 0, "brand": brand, "species": "GYPSOPHILA",
            })
            # Also index just by the paniculata variety name for simpler lookups
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": 0,
                "spb": 0, "brand": brand, "species": "GYPSOPHILA",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse carnations: CLAVEL [COL|EC] [GRADE] COLOR SIZECM SPBU [BRAND]
        m = re.match(
            r"(?:MINI\s+)?CLAVEL\s+(?:COL\s+|EC\s+|SPRAY\s+)?(?:FANCY\s+)?(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "CARNATIONS",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse hydrangeas: HYDRANGEA VARIETY SIZECM SPB [BRAND]
        m = re.match(
            r"HYDRANGEA\s+(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "HYDRANGEAS",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse alstroemeria
        m = re.match(
            r"ALSTRO?E?MERIA\s+(?:EC\s+|COL\s+)?(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "ALSTROEMERIA",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Parse chrysanthemum: CRISANTEMO VARIETY SIZECM SPBU [BRAND]
        m = re.match(
            r"CRISANTEMO\s+(.+?)\s+(\d+)CM\s+(\d+)U(?:\s+(.+))?$",
            n_upper,
        )
        if m:
            variety = m.group(1).strip()
            size = int(m.group(2))
            spb = int(m.group(3))
            brand = (m.group(4) or "").strip()
            variety_index[normalize(variety)].append({
                "id": art_id, "nombre": nombre, "size": size,
                "spb": spb, "brand": brand, "species": "CHRYSANTHEMUM",
            })
            if brand:
                brand_index[normalize(brand)].append((art_id, nombre))
            continue

        # Generic: extract last word as potential brand
        words = n_upper.split()
        if len(words) >= 2 and words[-1].isalpha() and len(words[-1]) > 2:
            brand_index[normalize(words[-1])].append((art_id, nombre))

    return variety_index, name_index, brand_index, provider_index


def get_brands_for_provider(provider_name: str) -> list[str]:
    """Get brand suffixes for a provider name."""
    return PROVIDER_BRAND_MAP.get(provider_name, [])


def translate_variety(variety: str, species: str) -> list[str]:
    """Generate translation variants for a variety name."""
    v = normalize(variety)
    variants = [v]

    # Check aliases
    if v in VARIETY_ALIASES:
        alias = VARIETY_ALIASES[v]
        if alias:
            variants.append(normalize(alias))

    # Strip "GARDEN ROSE " prefix
    if v.startswith("GARDEN ROSE "):
        variants.append(v.replace("GARDEN ROSE ", "").strip())

    # Translate individual words
    words = v.split()
    translated_words = []
    any_translated = False
    for w in words:
        if w in TRANSLATIONS:
            translated_words.append(TRANSLATIONS[w])
            any_translated = True
        else:
            translated_words.append(w)
    if any_translated:
        variants.append(" ".join(translated_words))

    # Handle GYPSOPHILA species → search as PANICULATA
    if species in ("GYPSOPHILA",) or "GYPSOPHILA" in v or "GYPSO" in v:
        # Strip "GYPSOPHILA " or "GYPSO. " prefix and translate rest
        cleaned = re.sub(r"^GYPSO(?:PHILA)?\.?\s*", "", v).strip()
        if cleaned:
            # Translate color words
            cleaned_words = cleaned.split()
            tr_words = [TRANSLATIONS.get(w, w) for w in cleaned_words]
            tr_cleaned = " ".join(tr_words)
            variants.append(f"PANICULATA {tr_cleaned}")
            variants.append(tr_cleaned)
            # Also try "OVERTIME" for "OVER TIME"
            if "OVER TIME" in cleaned:
                ot = cleaned.replace("OVER TIME", "OVERTIME")
                ot_words = ot.split()
                tr_ot = " ".join(TRANSLATIONS.get(w, w) for w in ot_words)
                variants.append(f"PANICULATA {tr_ot}")
                variants.append(tr_ot)

    # Handle TINTED → TEÑIDA
    if "TINTED" in v:
        rest = v.replace("TINTED", "").strip()
        rest_words = rest.split()
        tr_rest = " ".join(TRANSLATIONS.get(w, w) for w in rest_words)
        variants.append(f"TENIDA {tr_rest}")

    # Deduplicate while preserving order
    seen = set()
    result = []
    for x in variants:
        nx = normalize(x)
        if nx and nx not in seen:
            seen.add(nx)
            result.append(nx)
    return result


def fuzzy_match_variety(query: str, candidates: list[str], threshold: float = 0.75) -> list[tuple[str, float]]:
    """Fuzzy match a variety name against candidates."""
    query_n = normalize(query)
    results = []
    for c in candidates:
        ratio = SequenceMatcher(None, query_n, normalize(c)).ratio()
        if ratio >= threshold:
            results.append((c, ratio))
    return sorted(results, key=lambda x: -x[1])


def match_line(variety: str, talla, especie: str, proveedor: str,
               variety_index, name_index, brand_index, provider_index,
               articles, provider_id_map):
    """
    Match a single pending line following the 7-step process.

    Returns: (articulo_id, articulo_nombre, confianza, metodo)
    """
    variety_clean = normalize(variety)
    size = int(talla) if talla and str(talla).isdigit() else 0

    # Get brand variants for this provider
    brands = get_brands_for_provider(proveedor)
    brands_norm = [normalize(b) for b in brands]

    # Step 1+2: Search by variety + size
    variety_variants = translate_variety(variety, especie)

    all_candidates = []
    for v in variety_variants:
        if v in variety_index:
            all_candidates.extend(variety_index[v])

    # If no exact match, try fuzzy on variety index keys
    if not all_candidates:
        all_variety_keys = list(variety_index.keys())
        for v in variety_variants:
            fuzzy_results = fuzzy_match_variety(v, all_variety_keys, threshold=0.80)
            for fkey, ratio in fuzzy_results[:3]:
                for cand in variety_index[fkey]:
                    cand_copy = dict(cand)
                    cand_copy["fuzzy_ratio"] = ratio
                    all_candidates.append(cand_copy)

    if not all_candidates:
        return None, "", "SIN_MATCH", "no_candidatos"

    # Filter by size if available
    if size > 0:
        with_size = [c for c in all_candidates if c["size"] == size]
        if with_size:
            candidates = with_size
        else:
            candidates = all_candidates
    else:
        candidates = all_candidates

    # Step 3: PRIORITY — match by brand/farm in article name
    if brands_norm:
        with_brand = []
        for c in candidates:
            c_brand = normalize(c.get("brand", ""))
            if c_brand and any(b in c_brand or c_brand in b for b in brands_norm):
                with_brand.append(c)

        if with_brand:
            # Best match: brand + size
            best = with_brand[0]
            return best["id"], best["nombre"], "ALTA", f"variedad+talla+marca({best.get('brand', '')})"

    # Step 4: Match by provider ID in articles
    prov_id = provider_id_map.get(proveedor, 0)
    if prov_id:
        by_prov = [c for c in candidates if articles[c["id"]]["id_proveedor"] == prov_id]
        if by_prov:
            best = by_prov[0]
            return best["id"], best["nombre"], "MEDIA", f"variedad+talla+proveedor_id({prov_id})"

    # Step 5: Fallback to generic EC/COL
    generics = [c for c in candidates if not c.get("brand")]
    ec_col = [c for c in generics if "EC " in c["nombre"].upper() or "COL " in c["nombre"].upper()]
    if ec_col:
        best = ec_col[0]
        return best["id"], best["nombre"], "BAJA", "variedad+talla+generico_EC_COL"

    # Also try generic articles that have no brand
    if generics:
        best = generics[0]
        return best["id"], best["nombre"], "BAJA", "variedad+talla+sin_marca"

    # Step 6: Any candidate at all (most generic)
    if candidates:
        best = candidates[0]
        is_fuzzy = best.get("fuzzy_ratio", 0)
        if is_fuzzy:
            return best["id"], best["nombre"], "MUY_BAJA", f"fuzzy({is_fuzzy:.0%})"
        return best["id"], best["nombre"], "MUY_BAJA", "variedad_generica"

    # Step 7: No match
    return None, "", "SIN_MATCH", "sin_match"


def build_provider_id_map() -> dict:
    """Map provider name → VeraBuy provider ID from config."""
    from src.config import PROVIDERS
    name_to_id = {}
    for key, pdata in PROVIDERS.items():
        name_to_id[pdata["name"]] = pdata["id"]
    return name_to_id


def main():
    print("=" * 70)
    print("MATCHING MASIVO — Líneas Pendientes 2026")
    print("=" * 70)

    # Step A: Load data
    print("\n[1/5] Cargando artículos desde SQL...")
    articles = load_articles(SQL_FILE)
    print(f"  → {len(articles)} artículos cargados")

    print("[1/5] Construyendo índices...")
    variety_index, name_index, brand_index, provider_index = build_indices(articles)
    print(f"  → {len(variety_index)} variedades indexadas")
    print(f"  → {len(brand_index)} marcas indexadas")

    provider_id_map = build_provider_id_map()

    # Auto-discover additional brand mappings from articles
    print("[1/5] Descubriendo mapeos proveedor→marca desde artículos...")
    for prov_name, prov_id in provider_id_map.items():
        if prov_id == 0:
            continue
        if prov_name in PROVIDER_BRAND_MAP:
            continue
        # Check what brand suffix is most common for this provider
        arts_for_prov = provider_index.get(prov_id, [])
        if not arts_for_prov:
            continue
        suffix_counter = Counter()
        for aid, nombre in arts_for_prov:
            words = nombre.upper().split()
            if len(words) >= 2 and words[-1].isalpha() and len(words[-1]) > 2:
                suffix_counter[words[-1]] += 1
        if suffix_counter:
            top = suffix_counter.most_common(1)[0]
            if top[1] >= 3:
                PROVIDER_BRAND_MAP[prov_name] = [top[0]]
                print(f"  → Descubierto: {prov_name} → {top[0]} ({top[1]} artículos)")

    # Step B: Load pending lines
    print("\n[2/5] Cargando líneas pendientes...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Lineas Sin Match"]

    lines = []
    manual_lines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        pdf, prov, desc, especie, variedad, talla, tallos, codigo = row
        line_data = {
            "row": i + 1,
            "pdf": pdf,
            "proveedor": prov,
            "descripcion": desc,
            "especie": especie or "ROSES",
            "variedad": variedad or "",
            "talla": talla,
            "tallos": tallos,
            "codigo_manual": codigo if (codigo is not None and codigo != "" and codigo != 0) else None,
        }
        lines.append(line_data)
        if line_data["codigo_manual"] is not None:
            manual_lines.append(line_data)
    wb.close()

    print(f"  → {len(lines)} líneas totales")
    print(f"  → {len(manual_lines)} con código manual")
    print(f"  → {len(lines) - len(manual_lines)} pendientes de match")

    # Step C: Validate against manual mappings
    print("\n[3/5] Validando contra mappings manuales...")
    correct = 0
    wrong = 0
    wrong_details = []
    for ml in manual_lines:
        art_id, art_name, conf, method = match_line(
            ml["variedad"], ml["talla"], ml["especie"], ml["proveedor"],
            variety_index, name_index, brand_index, provider_index,
            articles, provider_id_map,
        )
        expected = int(ml["codigo_manual"])
        if art_id == expected:
            correct += 1
        else:
            wrong += 1
            wrong_details.append({
                "prov": ml["proveedor"],
                "variedad": ml["variedad"],
                "talla": ml["talla"],
                "esperado": expected,
                "esperado_nombre": articles.get(expected, {}).get("nombre", "?"),
                "obtenido": art_id,
                "obtenido_nombre": art_name,
                "metodo": method,
                "confianza": conf,
            })

    total_manual = len(manual_lines)
    rate = correct / total_manual * 100 if total_manual else 0
    print(f"\n  Validación: {correct}/{total_manual} correctos ({rate:.1f}%)")
    if wrong_details:
        print(f"  Errores ({wrong}):")
        for wd in wrong_details:
            print(f"    {wd['prov']:25s} | {wd['variedad']:25s} | {wd['talla']}")
            print(f"      Esperado:  {wd['esperado']} → {wd['esperado_nombre']}")
            print(f"      Obtenido:  {wd['obtenido']} → {wd['obtenido_nombre']} ({wd['metodo']})")

    if rate < 85:
        print(f"\n  ⚠ TASA DE VALIDACIÓN ({rate:.1f}%) < 85%. Revisar errores antes de aplicar.")
    else:
        print(f"\n  ✓ TASA DE VALIDACIÓN OK ({rate:.1f}% >= 85%)")

    # Step D: Run matching on all lines
    print("\n[4/5] Ejecutando matching masivo...")
    results = []
    counts = {"ALTA": 0, "MEDIA": 0, "BAJA": 0, "MUY_BAJA": 0, "SIN_MATCH": 0, "MANUAL": 0}

    for line in lines:
        if line["codigo_manual"] is not None:
            art = articles.get(int(line["codigo_manual"]), {})
            results.append({
                **line,
                "codigo_found": int(line["codigo_manual"]),
                "nombre_found": art.get("nombre", ""),
                "confianza": "MANUAL",
                "metodo": "manual_usuario",
            })
            counts["MANUAL"] += 1
            continue

        art_id, art_name, conf, method = match_line(
            line["variedad"], line["talla"], line["especie"], line["proveedor"],
            variety_index, name_index, brand_index, provider_index,
            articles, provider_id_map,
        )
        results.append({
            **line,
            "codigo_found": art_id,
            "nombre_found": art_name,
            "confianza": conf,
            "metodo": method,
        })
        counts[conf] += 1

    # Step E: Generate output
    print("\n[5/5] Generando resultados...")

    # Excel output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Matching Results"
    headers = [
        "PDF", "Proveedor", "Descripcion", "Especie", "Variedad", "Talla",
        "Tallos", "CODIGO MANUAL", "CODIGO ENCONTRADO", "NOMBRE ARTICULO",
        "CONFIANZA", "METODO",
    ]
    ws_out.append(headers)

    for r in results:
        ws_out.append([
            r["pdf"], r["proveedor"], r["descripcion"], r["especie"],
            r["variedad"], r["talla"], r["tallos"], r["codigo_manual"],
            r["codigo_found"], r["nombre_found"], r["confianza"], r["metodo"],
        ])

    wb_out.save(OUTPUT_EXCEL)
    print(f"  → Excel guardado: {OUTPUT_EXCEL}")

    # Synonyms
    syns_alta_media = {}
    syns_baja = {}
    for r in results:
        if r["codigo_found"] is None or r["confianza"] == "MANUAL":
            continue
        key = f"{r['proveedor']}|{r['especie']}|{r['variedad']}|{r['talla']}"
        entry = {
            "articulo_id": r["codigo_found"],
            "articulo_name": r["nombre_found"],
            "origen": "auto-matching",
            "provider": r["proveedor"],
            "species": r["especie"],
            "variety": r["variedad"],
            "size": r["talla"],
            "confianza": r["confianza"],
            "metodo": r["metodo"],
        }
        if r["confianza"] in ("ALTA", "MEDIA"):
            syns_alta_media[key] = entry
        else:
            syns_baja[key] = entry

    with open(OUTPUT_SYNS_ALTA, "w", encoding="utf-8") as f:
        json.dump(syns_alta_media, f, indent=2, ensure_ascii=False)
    print(f"  → Sinónimos ALTA/MEDIA: {len(syns_alta_media)} → {OUTPUT_SYNS_ALTA}")

    with open(OUTPUT_SYNS_BAJA, "w", encoding="utf-8") as f:
        json.dump(syns_baja, f, indent=2, ensure_ascii=False)
    print(f"  → Sinónimos BAJA/MUY_BAJA: {len(syns_baja)} → {OUTPUT_SYNS_BAJA}")

    # Summary report
    print("\n" + "=" * 70)
    print("INFORME RESUMEN")
    print("=" * 70)
    print(f"  Total líneas:                          {len(results)}")
    print(f"  Ya tenían código (manual):             {counts['MANUAL']}")
    print(f"  Match con marca (confianza ALTA):      {counts['ALTA']}")
    print(f"  Match por proveedor (confianza MEDIA): {counts['MEDIA']}")
    print(f"  Match genérico (confianza BAJA):       {counts['BAJA']}")
    print(f"  Match muy genérico (MUY BAJA):         {counts['MUY_BAJA']}")
    print(f"  Sin match:                             {counts['SIN_MATCH']}")
    print(f"")
    print(f"  Sinónimos nuevos ALTA/MEDIA:           {len(syns_alta_media)}")
    print(f"  Sinónimos revisión BAJA:               {len(syns_baja)}")
    print(f"  Validación manual:                     {correct}/{total_manual} ({rate:.1f}%)")
    print("=" * 70)

    # Per-provider breakdown
    prov_counts = defaultdict(lambda: {"total": 0, "matched": 0, "sin_match": 0})
    for r in results:
        p = r["proveedor"]
        prov_counts[p]["total"] += 1
        if r["codigo_found"] is not None:
            prov_counts[p]["matched"] += 1
        else:
            prov_counts[p]["sin_match"] += 1

    print("\nDesglose por proveedor:")
    print(f"  {'Proveedor':30s} | {'Total':>6s} | {'Match':>6s} | {'Sin':>6s} | {'%':>6s}")
    print("  " + "-" * 65)
    for p, c in sorted(prov_counts.items(), key=lambda x: -x[1]["total"]):
        pct = c["matched"] / c["total"] * 100 if c["total"] else 0
        print(f"  {p:30s} | {c['total']:6d} | {c['matched']:6d} | {c['sin_match']:6d} | {pct:5.1f}%")


if __name__ == "__main__":
    main()
