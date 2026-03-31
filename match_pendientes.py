"""Matching masivo de líneas pendientes contra artículos VeraBuy.

Usa el matcher mejorado del proyecto (src/articulos.py search_with_priority)
que replica el criterio manual del usuario:
  1. Variedad + talla + farm/marca en nombre artículo
  2. Variedad + talla + campo proveedor (id_proveedor)
  3. Variedad + talla + genérico EC/COL
  4. Variedad genérica sin talla
  5. Sin match → dejar vacío

Validación obligatoria contra los 88 mappings manuales antes de aplicar.
"""
from __future__ import annotations

import json
import logging
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from src.articulos import ArticulosLoader
from src.config import SQL_FILE, PROVIDERS

# --- Config ---
EXCEL_FILE = Path('lineas_pendientes_2026.xlsx')
OUTPUT_EXCEL = Path('lineas_pendientes_2026_matched.xlsx')
SYNS_HIGH_FILE = Path('sinonimos_auto_alta_media.json')
SYNS_LOW_FILE = Path('sinonimos_auto_baja_revision.json')

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')


# =============================================================================
# Mapeo proveedor nombre → provider_id (de PROVIDERS config)
# =============================================================================

def get_provider_id(name: str) -> int:
    """Busca el provider_id para un nombre de proveedor."""
    name_up = name.upper().strip()
    for pkey, pdata in PROVIDERS.items():
        if pdata['name'].upper() == name_up:
            return pdata['id']
    # Fuzzy: buscar por contenido
    for pkey, pdata in PROVIDERS.items():
        if name_up in pdata['name'].upper() or pdata['name'].upper() in name_up:
            return pdata['id']
    return 0


def get_provider_key(name: str) -> str:
    """Busca la clave del proveedor para un nombre."""
    name_up = name.upper().strip()
    for pkey, pdata in PROVIDERS.items():
        if pdata['name'].upper() == name_up:
            return pkey
    for pkey, pdata in PROVIDERS.items():
        if name_up in pdata['name'].upper() or pdata['name'].upper() in name_up:
            return pkey
    return ''


# =============================================================================
# Cargar y procesar Excel
# =============================================================================

def load_excel(path: Path) -> list[dict]:
    """Carga las líneas del Excel."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        rows.append({
            'pdf': vals[0] or '',
            'proveedor': vals[1] or '',
            'descripcion': vals[2] or '',
            'especie': vals[3] or '',
            'variedad': vals[4] or '',
            'talla': int(vals[5]) if vals[5] else 0,
            'tallos': vals[6],
            'codigo_manual': int(vals[7]) if vals[7] else None,
        })
    wb.close()
    return rows


def match_row(row: dict, art: ArticulosLoader) -> dict:
    """Matchea una fila usando search_with_priority del ArticulosLoader."""
    provider_id = get_provider_id(row['proveedor'])
    provider_key = get_provider_key(row['proveedor'])

    result, confidence, method = art.search_with_priority(
        variety=row['variedad'],
        size=row['talla'],
        provider_id=provider_id,
        provider_key=provider_key,
    )

    if result:
        row['codigo_found'] = result['id']
        row['nombre_found'] = result['nombre']
        row['confianza'] = confidence
        row['metodo'] = method
    else:
        row['codigo_found'] = None
        row['nombre_found'] = ''
        row['confianza'] = ''
        row['metodo'] = 'sin_match'

    return row


def validate_manual(rows: list[dict], art: ArticulosLoader) -> tuple[int, int, list]:
    """Valida el algoritmo contra los mappings manuales."""
    manual = [r for r in rows if r['codigo_manual']]
    correct = 0
    wrong = []

    for r in manual:
        provider_id = get_provider_id(r['proveedor'])
        provider_key = get_provider_key(r['proveedor'])

        result, confidence, method = art.search_with_priority(
            variety=r['variedad'],
            size=r['talla'],
            provider_id=provider_id,
            provider_key=provider_key,
        )

        found_id = result['id'] if result else None
        if found_id == r['codigo_manual']:
            correct += 1
        else:
            expected_name = art.articulos.get(r['codigo_manual'], {}).get('nombre', '?')
            wrong.append({
                'proveedor': r['proveedor'],
                'variedad': r['variedad'],
                'talla': r['talla'],
                'esperado': r['codigo_manual'],
                'obtenido': found_id,
                'nombre_esperado': expected_name,
                'nombre_obtenido': result['nombre'] if result else '',
                'metodo': method,
                'confianza': confidence,
            })

    return correct, len(manual), wrong


def save_excel(rows: list[dict], path: Path) -> None:
    """Guarda el Excel con los resultados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lineas Matched'
    headers = ['PDF', 'Proveedor', 'Descripcion', 'Especie', 'Variedad', 'Talla',
               'Tallos', 'CODIGO MANUAL', 'CODIGO ENCONTRADO', 'NOMBRE ARTICULO',
               'CONFIANZA', 'METODO']
    ws.append(headers)
    for r in rows:
        ws.append([
            r['pdf'], r['proveedor'], r['descripcion'], r['especie'],
            r['variedad'], r['talla'], r['tallos'], r['codigo_manual'],
            r.get('codigo_found'), r.get('nombre_found', ''),
            r.get('confianza', ''), r.get('metodo', ''),
        ])
    wb.save(path)
    logger.info(f"Excel guardado: {path}")


def save_synonyms(rows: list[dict]) -> None:
    """Guarda sinónimos separados por nivel de confianza."""
    high, low = {}, {}
    for r in rows:
        if not r.get('codigo_found') or r.get('confianza') == 'MANUAL':
            continue
        key = f"{r['proveedor']}|{r['especie']}|{r['variedad']}|{r['talla']}"
        entry = {
            'articulo_id': r['codigo_found'],
            'articulo_name': r.get('nombre_found', ''),
            'proveedor': r['proveedor'],
            'especie': r['especie'],
            'variedad': r['variedad'],
            'talla': r['talla'],
            'confianza': r['confianza'],
            'metodo': r['metodo'],
            'origen': 'auto-matching',
        }
        if r['confianza'] in ('ALTA', 'MEDIA'):
            high[key] = entry
        else:
            low[key] = entry

    with open(SYNS_HIGH_FILE, 'w', encoding='utf-8') as f:
        json.dump(high, f, indent=2, ensure_ascii=False)
    logger.info(f"Sinónimos ALTA/MEDIA: {len(high)} en {SYNS_HIGH_FILE}")

    with open(SYNS_LOW_FILE, 'w', encoding='utf-8') as f:
        json.dump(low, f, indent=2, ensure_ascii=False)
    logger.info(f"Sinónimos BAJA/MUY_BAJA para revisión: {len(low)} en {SYNS_LOW_FILE}")


def print_report(rows: list[dict]) -> None:
    """Imprime el informe resumen."""
    conf = Counter(r.get('confianza', '') for r in rows)
    total = len(rows)
    sin_match = conf.get('', 0)

    print("\n" + "=" * 60)
    print("  INFORME DE MATCHING - LINEAS PENDIENTES 2026")
    print("=" * 60)
    print(f"  Total lineas:                          {total}")
    print(f"  Ya tenian codigo (manual):             {conf.get('MANUAL', 0)}")
    print(f"  Match con marca (confianza ALTA):      {conf.get('ALTA', 0)}")
    print(f"  Match por proveedor (confianza MEDIA): {conf.get('MEDIA', 0)}")
    print(f"  Match generico EC/COL (confianza BAJA):{conf.get('BAJA', 0)}")
    print(f"  Match articulo generico (MUY BAJA):    {conf.get('MUY_BAJA', 0)}")
    print(f"  Sin match (revision manual):           {sin_match}")
    print("-" * 60)
    auto = total - sin_match - conf.get('MANUAL', 0)
    base = total - conf.get('MANUAL', 0)
    pct = auto / base * 100 if base else 0
    print(f"  Tasa de matching automatico:           {pct:.1f}%")
    print(f"  Sinonimos nuevos (ALTA+MEDIA):         {conf.get('ALTA', 0) + conf.get('MEDIA', 0)}")
    print("=" * 60)


def main():
    logger.info("=== MATCHING MASIVO DE LINEAS PENDIENTES ===")

    # Cargar artículos
    logger.info("Cargando articulos...")
    art = ArticulosLoader()
    art.load_from_sql(str(SQL_FILE))

    # Cargar Excel
    logger.info("Cargando Excel...")
    rows = load_excel(EXCEL_FILE)
    logger.info(f"Lineas cargadas: {len(rows)}")

    # VALIDACIÓN OBLIGATORIA contra mappings manuales
    logger.info("Validando contra 88 mappings manuales...")
    correct, total_manual, wrong = validate_manual(rows, art)
    pct = correct / total_manual * 100 if total_manual else 0
    print(f"\n  VALIDACION: {correct}/{total_manual} coinciden ({pct:.1f}%)")
    if wrong:
        print(f"  Errores ({len(wrong)}):")
        for w in wrong[:30]:
            print(f"    {w['proveedor']} | {w['variedad']} {w['talla']}cm")
            print(f"      Esperado: {w['esperado']} ({w['nombre_esperado']})")
            print(f"      Obtenido: {w['obtenido']} ({w['nombre_obtenido']}) [{w['metodo']}|{w['confianza']}]")

    if pct < 85:
        logger.warning(f"Tasa de validacion ({pct:.1f}%) inferior al 85%!")
    else:
        logger.info(f"Validacion OK: {pct:.1f}% >= 85%")

    # Matching masivo
    logger.info("Ejecutando matching masivo...")
    for i, row in enumerate(rows):
        if row['codigo_manual']:
            row['codigo_found'] = row['codigo_manual']
            row['nombre_found'] = art.articulos.get(row['codigo_manual'], {}).get('nombre', '')
            row['confianza'] = 'MANUAL'
            row['metodo'] = 'manual'
        else:
            match_row(row, art)
        if (i + 1) % 2000 == 0:
            logger.info(f"  Procesadas {i+1}/{len(rows)} lineas...")

    # Resultados
    save_excel(rows, OUTPUT_EXCEL)
    save_synonyms(rows)
    print_report(rows)


if __name__ == '__main__':
    main()
